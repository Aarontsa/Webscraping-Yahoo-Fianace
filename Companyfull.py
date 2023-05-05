import requests
import pandas as pd
import datetime
from bs4 import BeautifulSoup
from requests.exceptions import ConnectionError
from openpyxl import load_workbook, Workbook
# from openpyxl import Workbook
import datetime
import os
import random
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment

def company_information(company_list):
    name=company_list
    url = f'https://finance.yahoo.com/quote/{name}?p={name}'
    response = requests.get(url, headers={
        'User-Agent':  'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36'})
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')

    Market_Caps = soup.find_all(
        'div', {'class': 'C($tertiaryColor) Fz(12px)'})
    data = []

    for Market_Cap in Market_Caps:
        if Market_Cap.contents:
            data.append(str(Market_Cap.contents[0].text))

    # array = ['NasdaqGS - NasdaqGS Real Time Price. Currency in USD']
    string = data[0]

    # split the string into two parts based on the "Currency in" keyword
    split_string = string.split('Currency in')

    # extract the exchange name and currency code
    exchange_name = split_string[0].split('-')[0].strip()
    currency_code = split_string[1].strip()
    
    return exchange_name,currency_code

def summary(company_list):
    name=company_list
    url = f'https://finance.yahoo.com/quote/{name}?p={name}'
    response = requests.get(url, headers={
        'User-Agent':  'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36'})
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')

    today_date = datetime.datetime.now().strftime("%Y-%m-%d")
    title = soup.find('h1', {'class': 'D(ib) Fz(18px)'}).text
    Currency = soup.find(
        'div', {'class': 'C($tertiaryColor) Fz(12px)'}).find('span').text
    Day_Range1 = soup.find_all(
        'td', {'class': 'Ta(end) Fw(600) Lh(14px)'})[4].text
    Day_Range2 = Day_Range1.replace(",","")
    Day_Range3 = Day_Range2.split("-")
    Dnumber1 = Day_Range3[0].strip()
    Dnumber2 = Day_Range3[1].strip()
    Day_Range = (float(Dnumber1)+float(Dnumber2))/2
    oneyear_Range1 = soup.find_all(
        'td', {'class': 'Ta(end) Fw(600) Lh(14px)'})[5].text
    oneyear_Range2 = oneyear_Range1.replace(",","")
    oneyear_Range3 = oneyear_Range2.split("-")
    Onumber1 = oneyear_Range3[0].strip()
    Onumber2 = oneyear_Range3[1].strip()
    oneyear_Range = (float(Onumber1)+float(Onumber2))/2
    Market_Cap = soup.find_all(
        'td', {'class': 'Ta(end) Fw(600) Lh(14px)'})[8].text
    PE_ratio_string = soup.find_all(
        'td', {'class': 'Ta(end) Fw(600) Lh(14px)'})[10].text
    # PE_ratio = float(PE_ratio_string)
    if PE_ratio_string == 'N/A':
        PE_ratio = 0.0
    else:
        PE_ratio = float(PE_ratio_string)   
    
    return today_date, title, Currency, Day_Range, oneyear_Range, Market_Cap, PE_ratio


def balance(company_list):

    name=company_list
    url = f'https://finance.yahoo.com/quote/{name}/balance-sheet?p={name}'
    response = requests.get(url, headers={
        'User-Agent':  'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36'})
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')

    # print("balance")
    # to put the value
    indices = check_location_Balance_Sheet(html)
    indices_len_full,Total_cell_len_full = check_location_full_length_Balance_Sheet(html)
    column_Total=Total_cell_len_full/indices_len_full
    b1=int(indices[0]*column_Total-2)
    b2=int(indices[1]*column_Total-2)
    

    assets_string = soup.find_all(
        'div', {'class': 'Ta(c) Py(6px) Bxz(bb) BdB Bdc($seperatorColor) Miw(120px) Miw(100px)--pnclg D(tbc)'})[b1].text
    assets = float(assets_string.replace(",",""))/1000
    equity_string = soup.find_all(
        'div', {'class': 'Ta(c) Py(6px) Bxz(bb) BdB Bdc($seperatorColor) Miw(120px) Miw(100px)--pnclg D(tbc)'})[b2].text
    equity = float(equity_string.replace(",",""))/1000
  
    return assets, equity

def cash(company_list):
    name=company_list
    url = f'https://finance.yahoo.com/quote/{name}/cash-flow?p={name}'
    response = requests.get(url, headers={
        'User-Agent':  'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36'})
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')

    # print("cash")
    # to put the value
    indices = check_location_Cash_Flow(html)
    indices_len_full,Total_cell_len_full = check_location_full_length_Cash_Flow(html)
    column_Total=Total_cell_len_full/indices_len_full
    c1=int(indices[0]*column_Total-2)

    flow_cash_string = soup.find_all(
        'div', {'class': 'Ta(c) Py(6px) Bxz(bb) BdB Bdc($seperatorColor) Miw(120px) Miw(100px)--pnclg D(tbc)'})[c1].text
    flow_cash = float(flow_cash_string.replace(",",""))/1000
    # print(flow_cash_string)

    return flow_cash

def income(company_list):
    name=company_list
    
    url = f'https://finance.yahoo.com/quote/{name}/financials?p={name}'#{name},BOSCHLTD.NS
    # print(url)
    response = requests.get(url, headers={
        'User-Agent':  'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36'})
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')

    # print("income")
    # to put the value
    indices = check_location_Income_Statement(html)
    indices_len_full,Total_cell_len_full = check_location_full_length_Income_Statement(html)
    column_Total=Total_cell_len_full/indices_len_full
    i1=int(indices[0]*column_Total-2)
    i2=int(indices[1]*column_Total-2)
    i3=int(indices[2]*column_Total-2)
    # print(indices)
    # print(indices_len_full)
    # print(Total_cell_len_full)
    # print(column_Total)
    # print(i1)
    # print(i2)
    # print(i3)

    report_date_num = soup.find_all(
        'div', {'class': 'Ta(c) Py(6px) Bxz(bb) BdB Bdc($seperatorColor) Miw(120px) Miw(100px)--pnclg D(ib) Fw(b)'})[i1].text
    report_date = datetime.datetime.strptime(report_date_num, '%m/%d/%Y').strftime("%B %Y")
    rev_string = soup.find_all(
        'div', {'class': 'Ta(c) Py(6px) Bxz(bb) BdB Bdc($seperatorColor) Miw(120px) Miw(100px)--pnclg D(tbc)'})[i1].text
    rev = float(rev_string.replace(",",""))/1000
    pbt_string = soup.find_all(
        'div', {'class': 'Ta(c) Py(6px) Bxz(bb) BdB Bdc($seperatorColor) Miw(120px) Miw(100px)--pnclg D(tbc)'})[i2].text    
    pbt = float(pbt_string.replace(",",""))/1000
    ebitda_string = soup.find_all(
        'div', {'class': 'Ta(c) Py(6px) Bxz(bb) BdB Bdc($seperatorColor) Miw(120px) Miw(100px)--pnclg D(tbc)'})[i3].text    
    if ebitda_string == '-':
        ebitda = ebitda_string
    else:
        ebitda = float(ebitda_string.replace(",",""))/1000



    return report_date, rev, pbt, ebitda

def check_location_Income_Statement(html):
    #array indicator
    soup = BeautifulSoup(html, 'html.parser')
    Title_locations = soup.find_all('span', {'class': 'Va(m)'})
    data = []

    for Title_location in Title_locations:
        if Title_location.contents:
            data.append(str(Title_location.contents[0]))

    indices = []
    for item in ['Total Revenue', 'Pretax Income', 'EBITDA']:
        indices.append(data.index(item))

    # print("title match")
    # print(indices)
    return indices


def check_location_full_length_Income_Statement(html):
    # title column
    soup = BeautifulSoup(html, 'html.parser')
    Title_locations = soup.find_all('span', {'class': 'Va(m)'})
    data1_5 = []

    for Title_location in Title_locations:
        if Title_location.contents:
            data1_5.append(str(Title_location.contents[0]))

    indices_len_full = len(data1_5)-1
    # print("full title column")
    # print(indices_len_full)

    # whtie column 'div', {'class': 'Ta(c) Py(6px) Bxz(bb) BdB Bdc($seperatorColor) Miw(120px) Miw(100px)--pnclg D(tbc)'}
    soup = BeautifulSoup(html, 'html.parser')
    Total_cells = soup.find_all('div', {'class': 'Ta(c) Py(6px) Bxz(bb) BdB Bdc($seperatorColor) Miw(120px) Miw(100px)--pnclg D(tbc)'})
    data2_5 = []

    for Total_cell in Total_cells:
        if Total_cell.contents:
            data2_5.append(str(Total_cell.contents[0]))

    Total_cell_len_full = len(data2_5)
    # print("full cells")
    # print(Total_cell_len_full)

    return indices_len_full,Total_cell_len_full

def check_location_Balance_Sheet(html):
    #array indicator
    soup = BeautifulSoup(html, 'html.parser')
    Title_locations = soup.find_all('span', {'class': 'Va(m)'})
    data = []

    for Title_location in Title_locations:
        if Title_location.contents:
            data.append(str(Title_location.contents[0]))

    indices = []
    for item in ['Total Assets', 'Total Equity Gross Minority Interest']:
        indices.append(data.index(item))

    # print("title match")
    # print(indices)
    return indices


def check_location_full_length_Balance_Sheet(html):
    # title column
    soup = BeautifulSoup(html, 'html.parser')
    Title_locations = soup.find_all('span', {'class': 'Va(m)'})
    data1_5 = []

    for Title_location in Title_locations:
        if Title_location.contents:
            data1_5.append(str(Title_location.contents[0]))

    indices_len_full = len(data1_5)-1
    # print("full title column")
    # print(data1_5)

    # whtie column 'div', {'class': 'Ta(c) Py(6px) Bxz(bb) BdB Bdc($seperatorColor) Miw(120px) Miw(100px)--pnclg D(tbc)'}
    soup = BeautifulSoup(html, 'html.parser')
    Total_cells = soup.find_all('div', {'class': 'Ta(c) Py(6px) Bxz(bb) BdB Bdc($seperatorColor) Miw(120px) Miw(100px)--pnclg D(tbc)'})
    data2_5 = []

    for Total_cell in Total_cells:
        if Total_cell.contents:
            data2_5.append(str(Total_cell.contents[0]))

    Total_cell_len_full = len(data2_5)
    # print("full cells")
    # print(data2_5)

    return indices_len_full,Total_cell_len_full

def check_location_Cash_Flow(html):
    #array indicator
    soup = BeautifulSoup(html, 'html.parser')
    Title_locations = soup.find_all('span', {'class': 'Va(m)'})
    data = []

    for Title_location in Title_locations:
        if Title_location.contents:
            data.append(str(Title_location.contents[0]))

    indices = []
    for item in ['Free Cash Flow']:
        indices.append(data.index(item))

    # print("title match")
    # print(indices)
    return indices


def check_location_full_length_Cash_Flow(html):
    # title column
    soup = BeautifulSoup(html, 'html.parser')
    Title_locations = soup.find_all('span', {'class': 'Va(m)'})
    data1_5 = []

    for Title_location in Title_locations:
        if Title_location.contents:
            data1_5.append(str(Title_location.contents[0]))

    indices_len_full = len(data1_5)-1
    # print("full title column")
    # print(data1_5)

    # whtie column 'div', {'class': 'Ta(c) Py(6px) Bxz(bb) BdB Bdc($seperatorColor) Miw(120px) Miw(100px)--pnclg D(tbc)'}
    soup = BeautifulSoup(html, 'html.parser')
    Total_cells = soup.find_all('div', {'class': 'Ta(c) Py(6px) Bxz(bb) BdB Bdc($seperatorColor) Miw(120px) Miw(100px)--pnclg D(tbc)'})
    data2_5 = []

    for Total_cell in Total_cells:
        if Total_cell.contents:
            data2_5.append(str(Total_cell.contents[0]))

    Total_cell_len_full = len(data2_5)
    # print("full cells")
    # print(data2_5)

    return indices_len_full,Total_cell_len_full

def main_sub(inputs,inputs2):

    #create excel file
    username = os.getlogin()
    wb = Workbook() 
    filename = "WebsScraping-Company_Finance"+datetime.date.today().strftime("%Y%m%d")+"-"+str(random.randint(0, 10))+".xlsx"
    filepath = r"C:\\Users\\"+username+"\\Downloads\\" + filename +""
    # print(filepath)
    wb.save(filepath)

    #create excel sheets
    ws1 = wb.create_sheet("Sheet_A")
    ws1.title = "Customers"
    ws2 = wb.create_sheet("Sheet_B")
    ws2.title = "Competitors"
    wb.save(filepath)

    #load excel file
    wb = load_workbook(filepath) 

    #delete excel sheets
    del wb['Sheet']
    
    Typeofcompanys = [1, 2]
    
    for Typeofcompany in Typeofcompanys:
        if Typeofcompany == 1:
            company_lists=inputs
            # print(company_lists)
            Typesheet='Customers'
            # sheet = wb["Customers"]
        else:
            company_lists=inputs2    
            # print(company_lists)
            Typesheet='Competitors'
            # sheet = wb["Competitors"]


        
        # company_lists=inputs
        # print(company_lists)
        # Select sheet
        sheet = wb[Typesheet]

        # Assign Data
        sheet['A1'] = 'Excel creation date : ' + datetime.datetime.now().strftime("%Y-%m-%d") #summary7[0]
        sheet['C1']= f'Financial Data Crawling ({Typesheet})'
        #
        sheet['A3'] = 'Company information'
        sheet['A3'].alignment = Alignment(horizontal='center')
        sheet['A4'] = 'Company'
        sheet['A5'] = 'Stock exchange based'
        sheet['A6'] = 'Currency'
        sheet['A7'] = 'Forex rate (USD) FIXED'
        sheet['A8'] = 'Financial Reporting Date'
        #
        sheet['A10'] = 'Income Statement'
        sheet['A10'].alignment = Alignment(horizontal='center')
        sheet['A11'] = 'Total Revenue'
        sheet['A12'] = 'Pretax Income'
        sheet['A13'] = 'EBITDA'
        #
        sheet['A15'] = 'Balance sheet'
        sheet['A15'].alignment = Alignment(horizontal='center')
        sheet['A16'] = 'Total Assets'
        sheet['A17'] = 'Total Equity'
        #
        sheet['A19'] = 'Cash Flow'
        sheet['A19'].alignment = Alignment(horizontal='center')
        sheet['A20'] = 'Free Cash Flow'
        #
        sheet['A22'] = 'Market summary'
        sheet['A22'].alignment = Alignment(horizontal='center')
        sheet['A23'] = 'Day''s Range (midrange)'
        sheet['A24'] = '1 year range (midrange)'
        sheet['A25'] = 'Market Cap'
        sheet['A26'] = 'PE ratio'

        #add color background
        columns = range(1,11)
        for i in columns:
            sheet.cell(row=3, column=i).fill = PatternFill(start_color="6869ee", end_color="6869ee", fill_type = "solid")
        for i in columns:
            sheet.cell(row=10, column=i).fill = PatternFill(start_color="6869ee", end_color="6869ee", fill_type = "solid")
        for i in columns:
            sheet.cell(row=15, column=i).fill = PatternFill(start_color="6869ee", end_color="6869ee", fill_type = "solid")
        for i in columns:
            sheet.cell(row=19, column=i).fill = PatternFill(start_color="6869ee", end_color="6869ee", fill_type = "solid")
        for i in columns:
            sheet.cell(row=22, column=i).fill = PatternFill(start_color="6869ee", end_color="6869ee", fill_type = "solid")

        #bold text
        sheet['C1'].font = Font(bold=True, size=20)
        Rows = range(3,27)
        for i in Rows:
            sheet.cell(row=i, column=1).font = Font(bold=True)

        #coloums width
        sheet.column_dimensions['A'].width = 28
        sheet.column_dimensions['C'].width = 25    
        sheet.column_dimensions['E'].width = 25
        sheet.column_dimensions['G'].width = 25    
        sheet.column_dimensions['I'].width = 25

        j=0
        for i, company_list in enumerate(company_lists):
            print(company_list)
            
        # for company_list in company_lists:
            company_information2 = company_information(company_list)
            # print(company_information2) #exchange_name,currency_code
            if company_information2[1] == 'USD':
                # company_lists=inputs
                # print(company_lists)
                Currency_location= int(1)
                # print(Currency_location)
            elif company_information2[1] == 'CNY':
                # company_lists=inputs2    
                # print(company_lists)
                Currency_location= int(7)
                # print(Currency_location)
            elif company_information2[1] == 'INR':
                # company_lists=inputs2    
                # print(company_lists)
                Currency_location= int(80)
                # print(Currency_location)
            elif company_information2[1] == 'TWD':
                # company_lists=inputs2    
                # print(company_lists)
                Currency_location= int(30)
                # print(Currency_location)
            else:
                # company_lists=inputs2    
                # print(company_lists)
                Currency_location='5 Default'
                # print(Currency_location)

            #ADD PARAMETER TO PASS
            #get 7 value from summary
            summary7 = summary(company_list)
            #get 2 value from balance sheet
            balance2 =balance(company_list)
            #get 1 value from flow cash
            flow_cash1 = cash(company_list)
            #get 4 value from income
            income4 = income(company_list)
            # print(flow_cash1)
            #LOOP BY ADD 2 COLUMNS
            #------------------------------------------------------------company1
            # sheet['C4'] = summary7[1]
            sheet.cell(row=4, column=3+j).value = summary7[1]
            sheet['C5'] = 'NSE India'
            sheet.cell(row=5, column=3+j).value = company_information2[0]#'NSE India'
            # sheet['C5'].alignment = Alignment(horizontal='right')
            sheet.cell(row=5, column=3+j).alignment = Alignment(horizontal='right')
            # sheet['C6'] = 'INR'
            sheet.cell(row=6, column=3+j).value = company_information2[1]
            # sheet['C6'].alignment = Alignment(horizontal='right')
            sheet.cell(row=6, column=3+j).alignment = Alignment(horizontal='right')
            # sheet['C7'] = int(80)
            sheet.cell(row=7, column=3+j).value = Currency_location#int(80)
            # sheet['C8'] = income4[0]
            sheet.cell(row=8, column=3+j).value = income4[0]
            # sheet['C8'].alignment = Alignment(horizontal='right')
            sheet.cell(row=8, column=3+j).alignment = Alignment(horizontal='right')

            # sheet['C10'] = '\'000'
            sheet.cell(row=10, column=3+j).value = '\'000'
            # sheet['C10'].alignment = Alignment(horizontal='center')
            sheet.cell(row=10, column=3+j).alignment = Alignment(horizontal='center')
            # sheet['C11'] = int(income4[1])
            sheet.cell(row=11, column=3+j).value = int(income4[1])
            # sheet['C12'] = int(income4[2])
            sheet.cell(row=12, column=3+j).value = int(income4[2])
            # sheet['C13'] = income4[3]
            sheet.cell(row=13, column=3+j).value = income4[3]
            # sheet['C13'].alignment = Alignment(horizontal='right')
            sheet.cell(row=13, column=3+j).alignment = Alignment(horizontal='right')
            
            # sheet['C15'] = '\'000'
            sheet.cell(row=15, column=3+j).value = '\'000'
            # sheet['C15'].alignment = Alignment(horizontal='center')
            sheet.cell(row=15, column=3+j).alignment = Alignment(horizontal='center')
            # sheet['C16'] = int(balance2[0])
            sheet.cell(row=16, column=3+j).value = int(balance2[0])
            # sheet['C17'] = int(balance2[1])
            sheet.cell(row=17, column=3+j).value = int(balance2[1])

            # sheet['C19'] = '\'000'
            sheet.cell(row=19, column=3+j).value = '\'000'
            # sheet['C19'].alignment = Alignment(horizontal='center')
            sheet.cell(row=19, column=3+j).alignment = Alignment(horizontal='center')
            # sheet['C20'] = int(flow_cash1)
            sheet.cell(row=20, column=3+j).value = int(flow_cash1)

            # sheet['C23'] = summary7[3]
            sheet.cell(row=23, column=3+j).value = summary7[3]
            # sheet['C24'] = summary7[4]
            sheet.cell(row=24, column=3+j).value = summary7[4]
            # sheet['C25'] = summary7[5]
            sheet.cell(row=25, column=3+j).value = summary7[5]
            # sheet['C25'].alignment = Alignment(horizontal='right')
            sheet.cell(row=25, column=3+j).alignment = Alignment(horizontal='right')
            # sheet['C26'] = summary7[6]
            sheet.cell(row=26, column=3+j).value = summary7[6]

            j=j+2

    wb.save(filepath)

def main():
    # create an empty array to store the inputs
    inputs = []

    print("Customers Stock code: 'BOSCHLTD.NS','FLEX','KN','MSI'")
    # get input from the user and append to the array 3 times
    for i in range(3):
        user_input = input("Enter a Customer "+ str(i+1) +":")
        inputs.append(user_input)

    inputs2 = []
    print("Competitors Stock code: '002384.SZ','3037.TW','3044.TW','4958.TW'")
    # get input from the user and append to the array 3 times
    for i in range(3):
        user_input2 = input("Enter a Customer "+ str(i+1) +":")
        inputs2.append(user_input2)

    # print(inputs)
    main_sub(inputs,inputs2)
    print("Done!!")

if __name__ == '__main__':
    main()
    